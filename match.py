#!/usr/bin/env python3
"""
Compare a check-list of SharePoint URLs against a directory of "sensitive data"
files (.xlsx or .csv). For every URL in the check-list that also appears in a
sensitive file, group the match by department (the path segment after /sites/)
and write one output .xlsx per department.

Match basis : exact full URL (whitespace-trimmed, case-insensitive).
Grouping    : department only (category is recorded as a column).

Column names default to the constants below and can be overridden per run:
  --checklist-column        URL column in the check-list      (ObjectId)
  --sensitive-column        URL column in the sensitive files (FileUrl)
  --sensitive-extra-column  extra column carried to output    (LastModifiedTime)

Header matching is case-insensitive and whitespace-trimmed.

Usage:
    python match.py \
        --sensitive-dir ./sensitive \
        --checklist     ./checklist.xlsx \
        --out-dir       ./output \
        --glob "*.csv"
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# csv default field-size limit is too small for very long rows; raise it.
csv.field_size_limit(10 * 1024 * 1024)

# ---------------------------------------------------------------------------
# Column names. Change here, or override per run with the matching CLI flags.
# ---------------------------------------------------------------------------
CHECKLIST_COLUMN = "ObjectId"          # the file listing the URLs to confirm
SENSITIVE_COLUMN = "FileUrl"           # URL column in the scanned files
SENSITIVE_EXTRA_COLUMN = "LastModifiedTime"  # extra column carried to output
# ---------------------------------------------------------------------------

# Department = first path segment after /sites/ (or /teams/ for Teams sites).
DEPT_RE = re.compile(r"/(?:sites|teams)/([^/?#]+)", re.IGNORECASE)


def normalize(url: str) -> str:
    """Normalization used only for matching, never for display/output."""
    return url.strip().rstrip("/").lower()


def department_of(url: str) -> str:
    m = DEPT_RE.search(url)
    return m.group(1) if m else "_no_department"


def safe_filename(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]", "_", name) or "_unnamed"


def _clean(value):
    """Cell/field value -> trimmed string, or None if blank. The full value is
    kept as-is (URLs may legitimately contain spaces, e.g. 'Shared Documents')."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _xlsx_rows(path: Path):
    """Yield (header_list, data_row_iter) for each non-empty worksheet."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows()
            try:
                header = next(rows)
            except StopIteration:
                continue
            yield [h.value for h in header], (
                # carry both value and hyperlink target per cell; read-only
                # cells expose no .hyperlink, so getattr-guard it
                [(c.value, getattr(getattr(c, "hyperlink", None), "target", None))
                 for c in r]
                for r in rows
            )
    finally:
        wb.close()


def _csv_rows(path: Path):
    """Yield (header_list, data_row_iter) for a CSV (single logical sheet)."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(fh, dialect)
        try:
            header = next(reader)
        except StopIteration:
            return
        # materialize: file closes when this function returns
        yield header, ([(v, None) for v in row] for row in reader)


def read_records(path: Path, wanted: dict):
    """Yield dicts keyed by the logical names in `wanted` (name -> header).

    For .xlsx every worksheet is searched. After iterating, the columns that
    were located are exposed on `read_records.found_names` and all headers
    seen on `read_records.last_headers`. Raises ValueError for an unsupported
    extension; missing columns simply yield None for that key.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        sheets = _csv_rows(path)
    elif suffix in (".xlsx", ".xlsm"):
        sheets = _xlsx_rows(path)
    else:
        raise ValueError(
            f"{path.name}: unsupported extension '{path.suffix}' "
            f"(expected .csv, .xlsx, or .xlsm)"
        )

    want_lc = {name: hdr.strip().lower() for name, hdr in wanted.items()}
    read_records.last_headers = set()
    read_records.found_names = set()

    for header, data_rows in sheets:
        idx = {}
        for i, h in enumerate(header):
            if h is None:
                continue
            hs = str(h).strip()
            read_records.last_headers.add(hs)
            for name, target in want_lc.items():
                if hs.lower() == target:
                    idx[name] = i
        if not idx:
            continue
        read_records.found_names |= set(idx)
        for cells in data_rows:
            rec = {}
            for name, i in idx.items():
                if i < len(cells):
                    val, link = cells[i]
                    rec[name] = _clean(val) or _clean(link)
                else:
                    rec[name] = None
            yield rec


def iter_checklist_urls(path: Path, column: str):
    """Yield every URL in the check-list's URL column. Raises ValueError if
    the column is absent."""
    got = False
    for rec in read_records(path, {"url": column}):
        got = True
        if rec.get("url"):
            yield rec["url"]
    if not got and "url" not in read_records.found_names:
        raise ValueError(
            f"{path.name}: no column named '{column}'. "
            f"Headers seen: {sorted(read_records.last_headers) or '(none)'}"
        )


def iter_sensitive_records(path: Path, url_col: str, extra_col: str):
    """Yield (url, extra_value) from a sensitive file. Raises ValueError if the
    URL column is absent (whole file unusable). A missing extra column is
    tolerated: extra_value is None and `.extra_missing` is set True."""
    iter_sensitive_records.extra_missing = False
    saw_any = False
    for rec in read_records(path, {"url": url_col, "extra": extra_col}):
        saw_any = True
        if rec.get("url"):
            yield rec["url"], rec.get("extra")
    if "url" not in read_records.found_names and not saw_any:
        raise ValueError(
            f"{path.name}: no column named '{url_col}'. "
            f"Headers seen: {sorted(read_records.last_headers) or '(none)'}"
        )
    if "extra" not in read_records.found_names:
        iter_sensitive_records.extra_missing = True


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--sensitive-dir", required=True, type=Path,
                   help="Directory of sensitive files (secrets.csv, ...)")
    p.add_argument("--checklist", required=True, type=Path,
                   help="Single file with the URLs to confirm")
    p.add_argument("--out-dir", required=True, type=Path,
                   help="Where per-department .xlsx files are written")
    p.add_argument("--checklist-column", default=CHECKLIST_COLUMN,
                   help=f"URL column in the check-list (default: {CHECKLIST_COLUMN})")
    p.add_argument("--sensitive-column", default=SENSITIVE_COLUMN,
                   help=f"URL column in the sensitive files (default: {SENSITIVE_COLUMN})")
    p.add_argument("--sensitive-extra-column", default=SENSITIVE_EXTRA_COLUMN,
                   help="Extra column from the sensitive files carried into "
                        f"the output (default: {SENSITIVE_EXTRA_COLUMN})")
    p.add_argument("--glob", default="*.xlsx",
                   help="Filename pattern for sensitive files (default: *.xlsx)")
    args = p.parse_args()

    if not args.sensitive_dir.is_dir():
        p.error(f"--sensitive-dir not a directory: {args.sensitive_dir}")
    if not args.checklist.is_file():
        p.error(f"--checklist not found: {args.checklist}")

    # 1. Build the check-list lookup: normalized URL -> original URL.
    try:
        checklist = {}
        for url in iter_checklist_urls(args.checklist, args.checklist_column):
            checklist.setdefault(normalize(url), url)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1
    print(f"check-list: {len(checklist)} unique URLs from "
          f"{args.checklist.name} (column '{args.checklist_column}')")
    if not checklist:
        print("Nothing to match against; exiting.", file=sys.stderr)
        return 1

    sensitive_files = sorted(f for f in args.sensitive_dir.glob(args.glob)
                             if f.is_file() and not f.name.startswith("~$"))
    if not sensitive_files:
        print(f"No files matching {args.glob} in {args.sensitive_dir}",
              file=sys.stderr)
        return 1

    # 2. Scan each sensitive file; collect matches grouped by department.
    #    dept -> set of (url, extra, category, source_file)  [set = dedup]
    by_dept = defaultdict(set)
    print(f"scanning {len(sensitive_files)} sensitive file(s) "
          f"(url='{args.sensitive_column}', extra='{args.sensitive_extra_column}'):")
    for sf in sensitive_files:
        category = sf.stem
        matched_here = 0
        try:
            for url, extra in iter_sensitive_records(
                    sf, args.sensitive_column, args.sensitive_extra_column):
                key = normalize(url)
                if key in checklist:
                    by_dept[department_of(url)].add(
                        (url, extra or "", category, sf.name))
                    matched_here += 1
            note = f"matches={matched_here}"
            if iter_sensitive_records.extra_missing:
                note += f"  [warn: no '{args.sensitive_extra_column}' column]"
        except ValueError as e:
            note = f"SKIPPED ({e})"
        print(f"  {sf.name:<28} category={category:<14} {note}")

    if not by_dept:
        print("\nNo matches found. No output files written.")
        return 0

    # 3. Write one workbook per department.
    args.out_dir.mkdir(parents=True, exist_ok=True)
    print()
    grand_total = 0
    for dept in sorted(by_dept):
        rows = sorted(by_dept[dept], key=lambda r: (r[2], r[0]))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = safe_filename(dept)[:31] or "Sheet1"
        ws.append([args.sensitive_column, args.sensitive_extra_column,
                   "Sensitive Category", "Source File"])
        for url, extra, category, src in rows:
            ws.append([url, extra, category, src])
        ws.freeze_panes = "A2"
        out_path = args.out_dir / f"{safe_filename(dept)}.xlsx"
        wb.save(out_path)
        grand_total += len(rows)
        print(f"  {out_path.name:<28} {len(rows)} matched URL(s)")

    print(f"\nDone: {grand_total} match row(s) across "
          f"{len(by_dept)} department file(s) in {args.out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
